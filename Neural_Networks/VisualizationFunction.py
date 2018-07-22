"""
load_data - 
train_loss - function which calculates train loss
test_loss - function which calculates test loss
calculate_time - function which convert time of running program to minutes and seconds
epoch_progress - 
result_plot - plot the loss functions
save_as_mat - 
save_net_information - save all informations about neural net into .txt file
get_value - obtain hard to get informations from pytorch
plot_output - plot output images
net_specification - create dict with model sepcification
save_model - 
create_report -
"""

import os
#import shutil

import torch
from torch.autograd import Variable

#import time
import datetime

import numpy as np
import matplotlib.pyplot as plt
import matplotlib.font_manager as font_manager
from math import sqrt
#import scipy.io as sio  # to save mat file
from scipy.io import savemat
from openpyxl import load_workbook
from openpyxl import drawing

import warnings
warnings.filterwarnings("ignore")



# ============================= PREPARE DATA ==================================
def load_data(Inputs, Labels, split, batch_size, device, img_size):
    
    num_channels = 1
    
    if type(Inputs) == torch.Tensor:
        Inputs = Variable(Inputs)
    else:
        Inputs = Variable(torch.from_numpy(Inputs))
    Labels = Variable(torch.from_numpy(Labels))
    
    train_dataset = int((Inputs.size(0))*split)

    Train_Inputs = Inputs[ : train_dataset]
    Train_Labels = Labels[ : train_dataset]
    Test_Inputs = Inputs[train_dataset : ]
    Test_Labels = Labels[train_dataset : ]
    
    train_inputs = Variable(Train_Inputs.float()).to(device)
    train_labels = Variable(Train_Labels.float()).to(device)
    test_inputs = Variable(Test_Inputs.float()).to(device)
    test_labels = Variable(Test_Labels.float()).to(device)
    
    train_inputs = train_inputs.view(int(Train_Inputs.size(0)/batch_size), batch_size, num_channels, Inputs.size(-2), Inputs.size(-1))
    train_labels = train_labels.view(int(Train_Labels.size(0)/batch_size), batch_size, num_channels, Labels.size(-2), Labels.size(-1))
    test_inputs = test_inputs.view(int(Test_Inputs.size(0)/batch_size), batch_size, num_channels, Inputs.size(-2), Inputs.size(-1))
    test_labels = test_labels.view(int(Test_Labels.size(0)/batch_size), batch_size, num_channels, Labels.size(-2), Labels.size(-1))
    
    
    return train_inputs, train_labels, test_inputs, test_labels



    
# ============================= TRAIN LOSS ====================================
def train_loss(model, train_inputs, train_labels, mini_batch_size, criterion, device):
    
    num_channels = 1
    
    # prepare to size the size of the tensor
    train_inputs = train_inputs.view(train_inputs.size(0) * mini_batch_size, num_channels, train_inputs.size(-2), train_inputs.size(-1))
    train_labels = train_labels.view(train_labels.size(0) * mini_batch_size, num_channels, train_labels.size(-2), train_labels.size(-1))
    
    # obtain prediction
    train_outputs = model(train_inputs)
        
    # validate the size of the output
    train_outputs = train_outputs.view(train_labels.size(0), num_channels, train_labels.size(-2), train_labels.size(-1))
    
    # calculate test_loss
    train_loss = criterion(train_outputs, train_labels).cpu().data.numpy()
    train_loss = np.round(float(train_loss), 4)
    
    # change to view the same as input files
    train_outputs = train_outputs.view(int(train_labels.size(0)/mini_batch_size), mini_batch_size, num_channels, train_labels.size(-2), train_labels.size(-1))

    
    return train_loss, train_labels, train_outputs



# ============================ TEST LOSS ======================================
def test_loss(model, test_inputs, test_labels, mini_batch_size, criterion, device):
    
    num_channels = 1
    
    # prepare to size the size of the tensor
    test_inputs = test_inputs.view(test_inputs.size(0) * mini_batch_size, num_channels, test_inputs.size(-2), test_inputs.size(-1))
    test_labels = test_labels.view(test_labels.size(0) * mini_batch_size, num_channels, test_labels.size(-2), test_labels.size(-1))

    # obtain prediction
    test_outputs = model(test_inputs)
    
    # validate the size of the output
    test_outputs = test_outputs.view(test_labels.size(0), num_channels, test_labels.size(-2), test_labels.size(-1))
    
    # calculate test_loss
    test_loss = criterion(test_outputs, test_labels).cpu().data.numpy()
    test_loss = np.round(float(test_loss), 4)
    
    # change to view the same as input files
    test_outputs = test_outputs.view(int(test_labels.size(0)/mini_batch_size), mini_batch_size, num_channels, test_labels.size(-2), test_labels.size(-1))

    
    return test_loss, test_labels, test_outputs


# ============================= CALCULATE MIN / SEC ===========================
def calculate_time(start_time, end_time):
    # ------------ calculate time ---------------
    # place at top:         start = time.time()
    # place in the loop:    end = time.time()
    time_diff = round(float(end_time - start_time), 2)
    time_min = int(time_diff/60)
    time_sec = round(time_diff - 60*time_min)

    return time_min, time_sec


# =========================== IMAGE DIFFERNECE ================================
def image_difference(test_label, test_output):
    
    img_pixels = test_label.size(3) * test_label.size(4)
    test_label = test_label.view(test_label.size(0)*test_label.size(1),test_label.size(3), test_label.size(4))
    test_output = test_output.view(test_output.size(0)*test_output.size(1),test_output.size(3), test_output.size(4))
    img_diff = 0
    
    for i in range(test_output.size(0)):
        temp_test_label = test_label[i,:,:]
        temp_test_output = test_output[i,:,:]
        temp_test_label = temp_test_label.cpu().data.numpy()
        temp_test_output = temp_test_output.cpu().data.numpy()
        matrix = temp_test_label - temp_test_output
        img_diff += sqrt((matrix.sum())**2) / img_pixels
    
    total_error = img_diff / test_output.size(0)
    total_error = np.round(float(total_error), 5)

    return total_error


# =========================== PRINT PROGRESS ==================================
def epoch_progress(epoch, num_epoch, mini_batch_size, model, train_input, train_output, test_input,
                   test_output, criterion, start_time, end_time, img_size, device):
    
    # obtain train_loss
    train_loss_value, train_label, train_output= train_loss(model, train_input, train_output, mini_batch_size, criterion, device)
    
    # obtain test_loss
    test_loss_value, test_label, test_output = test_loss(model, test_input, test_output, mini_batch_size, criterion, device)
    
    # calculate time
    epoch_time_min, epoch_time_sec = calculate_time(start_time, end_time)
    
    # store data of the epoch
    store_data = np.zeros((1, 3))
    store_data[0, 0] = epoch
    store_data[0, 1] = train_loss_value
    store_data[0, 2] = test_loss_value

    
    # print progress
    print("{}/{} ||  train_loss: {}  |  test_loss: {}   |   {} min {} s".format(epoch,
              num_epoch, train_loss_value, test_loss_value, epoch_time_min, epoch_time_sec))

    return store_data, train_output, test_output




# =========================== PLOT ACCURACY ===================================
def result_plot(stage, all_store_data, criterion, report=False, saveJPG=False,
                saveSVG=False, savePDF=False):

    
    # font definition
    title_font = {'fontname':'Arial', 'size':'16', 'color':'black', 'weight':'normal',
                  'verticalalignment':'bottom'} # Bottom vertical alignment for more space
    axis_font = {'fontname':'Arial', 'size':'14'}
    font_prop = font_manager.FontProperties(size=11)
    
    # define figure
    fig, ax = plt.subplots(figsize=(12,5), ncols=2, nrows=1)
    
    left   =  0.1    # the left side of the subplots of the figure
    right  =  1.1    # the right side of the subplots of the figure
    bottom =  0.2    # the bottom of the subplots of the figure
    top    =  0.8    # the top of the subplots of the figure
    wspace =  .3    # the amount of width reserved for blank space between subplots
    hspace =  .1     # the amount of height reserved for white space between subplots


    # subplot 1
    ax[0].hold(True)
    ax[0].plot(all_store_data[:,0], all_store_data[:,1], color='#2A6F92')
    ax[0].plot(all_store_data[:,0], all_store_data[:,2], color='#DA9124')
    ax[0].set_xlabel('epochs', **axis_font)
    ax[0].set_ylabel('loss', **axis_font)
    ax[0].set_title('train / test loss', **title_font)
    ax[0].legend(['train loss', 'test loss'], prop=font_prop, numpoints=1)
    ax[0].hold(False)
    
    # subplot 2
    ax[1].plot(all_store_data[:,0], all_store_data[:,1], color='#2A6F92', label='test loss')
    ax[1].set_xlabel('epochs', **axis_font)
    ax[1].set_title('test loss', **title_font)
    
    
    if report == True:
        plt.subplots_adjust(wspace = wspace, bottom = bottom, hspace  =  hspace)
        w = int(25 / 2.5)  # 30
        h = int(8 / 2.5)  # 10
        fig.set_size_inches(w,h)
        fig.savefig('report_plot.png', dpi=85)
        
    
    # adjust image
    w = int(30 / 2.5)  # 30
    h = int(10 / 2.5)  # 10
    fig.set_size_inches(w,h)
    
    font_prop = font_manager.FontProperties(size=12)
    ax[0].legend(['train loss', 'test loss'], prop=font_prop, numpoints=1)
    
    top    =  0.8
    plt.subplots_adjust(top = top, hspace = hspace)

    
    
    # define title
    plot_title = "Loss values of {} | epochs: {} | {}".format(stage, len(all_store_data), criterion)
    fig.suptitle(plot_title, fontsize=16)
    
    
    # -------------------- save the plot ------------------------
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    short_title = "{}_{}_{}".format(stage, len(all_store_data), time_data)

    
    # create empty folder
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')

    if saveJPG == True:            
        fig.savefig('./RESULTS/' + short_title + ".jpg")
    if saveSVG == True:  
        fig.savefig('./RESULTS/' + short_title + ".svg")
    if savePDF == True:  
        fig.savefig('./RESULTS/' + short_title + ".pdf")
            
    



# =========================== SAVE AS .MAT ====================================
def save_as_mat(name, input_array, stage, num_epoch):
    
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    short_title = "{}_{}_{}".format(stage, num_epoch, time_data)
    
    # create empty folder
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')
    
    # save .mat file
    savemat('./RESULTS/' + name + '_' + short_title + '.mat', input_array)


# =========================== SAVE ARRAY ======================================
# save model information to text file.

def save_net_information(stage, name, model, model_specification):    
    
    # file name
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    
    # add title
    temp_model = []
    temp_model.append(name)
    temp_model.append(time_data)
    
    # space
    for m in range(4):
        temp_model.append('\n')
    
    # add model
    temp_model.append('------------------------- MODEL STRUCTURE ----------------------')
    temp_model.append('\n')
    temp_model.append('\n')
    list_model = get_value(model)
    for i in range(len(list_model)):
        temp_model.append(str(list_model[i]))
        temp_model.append('\n')
    temp_model.append('\n')
    
    # space
    for m in range(3):
        temp_model.append('\n')
    
    # add net specification
    temp_model.append('---------------------- MODEL SPECIFICATION ---------------------')
    temp_model.append('\n')
    for key, value in model_specification.items():
        spec_key = key
        spec_value = value
        temp_model.append(str(spec_key) + ': ' + str(spec_value))
        temp_model.append('\n')
    temp_model.append('\n')
    
    # create empty folder
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')
    
    # save file
    file = open('./RESULTS/' + name + '.txt', 'x')
    file.writelines(temp_model)
    file.close()



# ======================= GET VALUES FROM PYTORCH =============================
# function that takes the files from pytorch that are difficult to obtain
# e.g. model, and convert them into list."""

def get_value(input):
    
    with open('temp_file.txt', 'w') as f:
        print(input, file=f)

    file = open('temp_file.txt', "r")
    lines = list(file)
    file.close()
    
    os.remove('temp_file.txt')
    
    return lines



# ========================= UPLOAD PLOTS ======================================
def plot_output(stage, epochs, test_labels, test_outputs, img_size, report=False, saveJPG=False,
                saveSVG=False, savePDF=False):
    
    output = test_labels.view(test_labels.size(0)*test_labels.size(1), test_labels.size(-2), test_labels.size(-1))
    pred_output = test_outputs.view(test_outputs.size(0)*test_outputs.size(1), test_outputs.size(-2), test_outputs.size(-1))
    
    data_numbers = [1, 2, 3, 4, 5]
    
    fig, ax = plt.subplots(figsize=(12,5), ncols=len(data_numbers), nrows=2)

    for i in range(len(data_numbers)):
        
        # output
        out = output[data_numbers[i], :, :]
        temp_out = (Variable(out).data).cpu().numpy()
        ax[0][i].imshow(temp_out.T, extent=(0, img_size, 0, img_size), origin='1')
        ax[0][i].set_axis_off()
        ax[0][i].set_title('output {}'.format(data_numbers[i]))
        
        # predicted output
        pred_out = pred_output[data_numbers[i], :, :]
        temp_pred_out = (Variable(pred_out).data).cpu().numpy()
        ax[1][i].imshow(temp_pred_out.T, extent=(0, img_size, 0, img_size), origin='1')
        ax[1][i].set_axis_off()
        ax[1][i].set_title('predicted output {}'.format(data_numbers[i]))
        
    
    if report == True:
        w = int(25 / 2.5)  # 30
        h = int(11 / 2.5)  # 10
        fig.set_size_inches(w,h)
        plt.tight_layout(pad=0.4, w_pad=0.5, h_pad=1.0)
        fig.savefig('report_output.png', dpi=85)
    
    # adjust image
    w = int(26 / 2.5)  # 30
    h = int(10 / 2.5)  # 10
    fig.set_size_inches(w,h)
    
    plt.tight_layout(pad=0.4, w_pad=0.5, h_pad=1.0)
    plt.subplots_adjust(top = 0.8)


    # define title
    plot_title = "Output of {} | epochs: {}".format(stage, epochs)
    fig.suptitle(plot_title, fontsize=16)
    
    
    # -------------------- save the plot ------------------------
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    short_title = "{}_{}_{}".format(stage, epochs, time_data)

    
    # create empty folder
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')

    if saveJPG == True:            
        fig.savefig('./RESULTS/' + short_title + ".jpg")
    if saveSVG == True:  
        fig.savefig('./RESULTS/' + short_title + ".svg")
    if savePDF == True:  
        fig.savefig('./RESULTS/' + short_title + ".pdf")
        


# ========================= UPLOAD IMAGES =====================================
def net_specification(stage, dataset_size, input_size, output_size, epochs,
                      mini_batch_size, learning_rate, momentum, loss_function,
                      optimizer, train_test_split, device, min_train_loss,
                      max_train_loss, min_test_loss, max_test_loss, train_accuracy,
                      test_accuracy, overall_run_time, accuracy, img_diff):
    
    # get optimizer
    optimizer_type = str(type(optimizer))
    optimizer_type = optimizer_type[20:]
    optimizer_type = optimizer_type[:-2]
    
    # round numbers
    train_accuracy = np.round(float(train_accuracy), 4)
    test_accuracy = np.round(float(test_accuracy), 4)
    
    # get time
    time_min, time_sec = calculate_time(overall_run_time[0], overall_run_time[1])
    all_time = str(time_min) + ' min  ' + str(time_sec) + ' sec'
    
    model_specification = dict(stage = stage,
                               dataset_size = dataset_size,
                               input_size = input_size,
                               output_size = output_size,
                               epochs = epochs,
                               mini_batch_size = mini_batch_size,
                               learning_rate = learning_rate,
                               momentum = momentum,
                               loss_function = loss_function,
                               optimizer = optimizer_type,
                               train_test_split = train_test_split,
                               device = device,
                               min_train_loss = min_train_loss,
                               max_train_loss = max_train_loss,
                               min_test_loss = min_test_loss,
                               max_test_loss = max_test_loss,
                               train_accuracy = train_accuracy,
                               test_accuracy = test_accuracy,
                               overall_run_time = all_time,
                               accuracy = accuracy,
                               img_diff = img_diff)
    
    # save as .mat file
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')
    
    name = 'model_spec'
    save_as_mat(name, model_specification, stage, epochs)
    
    
    return model_specification



# ========================== SAVE MODEL =======================================
def save_model(model, name):
    
    # create empty folder for model
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')

    torch.save(model, './RESULTS/' + name + '.pt')
    torch.save(model.state_dict(), './RESULTS/' + name + '_state.pt')
    


# ========================= CREATE REPORT =====================================
def create_report(model, stage, model_specification):
    
    # ----------------------- Read File ---------------------------------------
    # Load in the workbook
    wb = load_workbook('./template_test.xlsx')
    
    # grab the active worksheet
    ws = wb.active
    
    # file name
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    file_name = 'REPORT_' + stage + '_' + time_data
    
    # save .txt file
    save_net_information(stage, file_name, model, model_specification)
    
    # ----------------------- Upload Data to File -----------------------------
    # get model
    model_description = get_value(model)
    
    # upload model
    ws['A7'] = "".join(str(x) for x in model_description)
    
    # upload stage
    ws['D3'] = str(stage)
    
    # upload save specification
    ws['J8'] = str(model_specification['dataset_size'])
    ws['J9'] = str(model_specification['input_size'])
    ws['J10'] = str(model_specification['output_size'])
    ws['J11'] = str(model_specification['epochs'])
    ws['J12'] = str(model_specification['learning_rate'])
    ws['J13'] = str(model_specification['momentum'])
    ws['J14'] = str(model_specification['loss_function'])
    ws['J15'] = str(model_specification['optimizer'])
    ws['J16'] = str(model_specification['train_test_split'])
    ws['J17'] = str(model_specification['device'])
    ws['J18'] = str(model_specification['mini_batch_size'])
    
    ws['N8'] = str(model_specification['min_train_loss'])
    ws['N9'] = str(model_specification['max_train_loss'])
    ws['N10'] = str(model_specification['min_test_loss'])
    ws['N11'] = str(model_specification['max_test_loss'])
    ws['N12'] = str(model_specification['train_accuracy'])
    ws['N13'] = str(model_specification['test_accuracy'])
    ws['N17'] = str(model_specification['overall_run_time'])
#    ws['N18'] = str(model_specification['accuracy'])
    ws['N18'] = str(model_specification['img_diff'])
    
    # uplaod plot
    img = drawing.image.Image('report_plot.png')
    ws.add_image(img, 'A35')
    
    # uplaod output
    img = drawing.image.Image('report_output.png')
    ws.add_image(img, 'A49')
    
    # ------------------------Save File ---------------------------------------
    # save as .xlsx file
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')
        
    wb.save('./RESULTS/' + file_name + '.xlsx')
           
    os.system('cd RESULTS && convert ' + file_name + '.xlsx ' + file_name + '.pdf')
    os.system('mv  -v ./RESULTS/_dirname/magick-*.pdf ./RESULTS/' + file_name + '.pdf')
    os.system('cd RESULTS && rmdir _dirname && cd .. && rm report_plot.png && rm report_output.png')


        
        
        
        
