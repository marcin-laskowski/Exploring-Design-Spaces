"""
load_data - function that load and prepare input data
load_params - function that load and prepare input parameters
get_loss - function which calculates loss
calculate_time - function which convert time of running program to minutes and seconds
image_difference - caluclates difference between label and output image
epoch_progress - calcuates progress of training the neural network
result_plot - plot the loss functions
save_as_mat - save matrix as .mat file
save_net_information - save all informations about neural net into .txt file
get_value - obtain hard to get informations from pytorch
save_data - save matrix as .npy file
plot_output - plot output images
net_specification - create dict with model sepcification
save_model - save model and state_dict of the network
max_stress - get maximum stress of each picture in test_labels and test_outputs
plot_max_stress - plot difference between test_labels and test_outputs of the max stress
take_stress_values - get stress propagation of the node with maximum stress
draw_stress_propagation_plot - creates fours plots with differenece of the stress propagation between test_labels and test_outputs
create_report - creates first page of the final report
create_report_page2 - creates second page of the final report wih more detail inforamtion
plot_sample_param_and_label - function that allows visualize very fast params and labels
X_and_Y_channels - creates X and Y channels for the input image
mean_and_CI - creates plot with mean and standard deviation of the stress propagation
get_best_model - compare loss and save model for smallest loss value
get_loss_and_output - calculate loss and return outputs data
"""

import os
#import shutil

import torch
from torch.autograd import Variable

import datetime

import numpy as np
import matplotlib.pyplot as plt
plt.switch_backend('agg')   # for ssh 
import matplotlib.font_manager as font_manager
from mpl_toolkits.axes_grid1 import make_axes_locatable
from math import sqrt
#import scipy.io as sio  # to save mat file
from scipy.io import savemat
from openpyxl import load_workbook
from openpyxl import drawing
from PIL import Image, ImageDraw
import random

import warnings
warnings.filterwarnings("ignore")



# ============================= PREPARE DATA ==================================
def load_data(Inputs, Labels, split, batch_size, device, img_size):

    num_channels = 1

    if type(Inputs) == torch.Tensor:
        pass
    else:
        Inputs = torch.from_numpy(Inputs)
    
    if type(Labels) == torch.Tensor:
        pass
    else:
        Labels = torch.from_numpy(Labels)


    train_dataset = int((Inputs.size(0))*split)

    Train_Inputs = Inputs[: train_dataset]
    Train_Labels = Labels[: train_dataset]
    Test_Inputs = Inputs[train_dataset :]
    Test_Labels = Labels[train_dataset :]

    train_inputs = Variable(Train_Inputs.float()).to(device)
    train_labels = Variable(Train_Labels.float()).to(device)
    test_inputs = Variable(Test_Inputs.float()).to(device)
    test_labels = Variable(Test_Labels.float()).to(device)

    train_inputs = train_inputs.view(Train_Inputs.size(0), num_channels, Inputs.size(3), Inputs.size(4))
    train_labels = train_labels.view(Train_Labels.size(0), num_channels, Labels.size(3), Labels.size(4))
    test_inputs = test_inputs.view(Test_Inputs.size(0), num_channels, Inputs.size(3), Inputs.size(4))
    test_labels = test_labels.view(Test_Labels.size(0), num_channels, Labels.size(3), Labels.size(4))


    return train_inputs, train_labels, test_inputs, test_labels



# ==================== PREPARE DATA - ONLY PARAMETERS =========================
def load_params(Params, split, batch_size, device, img_size):

    num_channels = 1

    train_dataset = int((Params.size(0))*split)

    Train_Params = Params[: train_dataset]
    Test_Params = Params[train_dataset :]

    train_params = Variable(Train_Params.float()).to(device)
    test_params = Variable(Test_Params.float()).to(device)

    train_params = train_params.view(Train_Params.size(0), num_channels, Params.size(3), Params.size(4))
    test_params = test_params.view(Test_Params.size(0), num_channels, Params.size(3), Params.size(4))


    return train_params, test_params


# ============================ GET LOSS ======================================
def get_loss(model, inputs, labels, mini_batch_size, criterion, device):
    
    num_channels = inputs.size(-3)

    # prepare to size the size of the tensor
    inputs = inputs.view(inputs.size(0), num_channels, inputs.size(-2), inputs.size(-1))
    labels = labels.view(labels.size(0), num_channels, labels.size(-2), labels.size(-1))

    # obtain prediction
    n = 0
    loss = 0
    dataset = inputs.size(0)
    step = mini_batch_size
    
#    outputs = torch.zeros((dataset, labels.size(1), labels.size(2), labels.size(3)))

    for i in range(0, dataset, step):
        input = inputs[i:i+step, :, :, :]  # train_input.size --> (:, 1, 64, 64)
        label = labels[i:i+step, :, :, :]  # train_labels.size --> (:, 1, 64, 64)
        output = model(input)

        # calculate test_loss
        loss += criterion(output, label).cpu().data.numpy()

#        outputs[i:i+step, :, :, :] = output
        n += 1

#    total_loss = (loss / n) / dataset**2
    total_loss = loss / dataset**2
    total_loss = np.round(float(total_loss), 5)

    return total_loss



# ============================= CALCULATE MIN / SEC ===========================
def calculate_time(start_time, end_time):
    # ------------ calculate time ---------------
    # place at top:         start = time.time()
    # place in the loop:    end = time.time()
    time_diff = round(float(end_time - start_time), 2)
    time_min = int(time_diff/60)
    time_sec = round(time_diff - 60*time_min)

    return time_min, time_sec


# =========================== IMAGE DIFFERNECE ================================
def image_difference(test_label, test_output):

    img_pixels = test_label.size(2) * test_label.size(3)
    test_label = test_label.view(test_label.size(0)*test_label.size(1),test_label.size(-2), test_label.size(-1))
    test_output = test_output.view(test_output.size(0)*test_output.size(1),test_output.size(-2), test_output.size(-1))
    img_diff = 0

    for i in range(test_output.size(0)):
        temp_test_label = test_label[i,:,:]
        temp_test_output = test_output[i,:,:]
        temp_test_label = temp_test_label.cpu().data.numpy()
        temp_test_output = temp_test_output.cpu().data.numpy()
        matrix = temp_test_label - temp_test_output
        img_diff += sqrt((matrix.sum())**2) / img_pixels

    total_error = img_diff / test_output.size(0)
    total_error = np.round(float(total_error), 5)

    return total_error


# =========================== PRINT PROGRESS ==================================
def epoch_progress(epoch, num_epoch, mini_batch_size, model, train_inputs, train_labels, test_inputs,
                   test_labels, criterion, start_time, end_time, img_size, device):

    # obtain train_loss
    train_loss_value = get_loss(model, train_inputs, train_labels, mini_batch_size, criterion, device)

    # obtain test_loss
    test_loss_value = get_loss(model, test_inputs, test_labels, mini_batch_size, criterion, device)

    # calculate time
    epoch_time_min, epoch_time_sec = calculate_time(start_time, end_time)

    # store data of the epoch
    store_data = np.zeros((1, 3))
    store_data[0, 0] = epoch
    store_data[0, 1] = train_loss_value
    store_data[0, 2] = test_loss_value


    # print progress
    print("{}/{} ||  train_loss: {}  |  test_loss: {}   |   {} min {} s".format(epoch,
              num_epoch, train_loss_value, test_loss_value, epoch_time_min, epoch_time_sec))

    return store_data




# =========================== PLOT ACCURACY ===================================
def result_plot(stage, all_store_data, criterion, report=False, saveJPG=False,
                saveSVG=False, savePDF=False):


    # font definition
    title_font = {'fontname':'Arial', 'size':'16', 'color':'black', 'weight':'normal',
                  'verticalalignment':'bottom'} # Bottom vertical alignment for more space
    axis_font = {'fontname':'Arial', 'size':'14'}
    font_prop = font_manager.FontProperties(size=11)

    # define figure
    fig, ax = plt.subplots(figsize=(12,5), ncols=2, nrows=1)

    left   =  0.1    # the left side of the subplots of the figure
    right  =  1.1    # the right side of the subplots of the figure
    bottom =  0.2    # the bottom of the subplots of the figure
    top    =  0.8    # the top of the subplots of the figure
    wspace =  .3    # the amount of width reserved for blank space between subplots
    hspace =  .1     # the amount of height reserved for white space between subplots


    # subplot 1
    ax[0].hold(True)
    ax[0].plot(all_store_data[:,0], all_store_data[:,1], color='#2A6F92')
    ax[0].plot(all_store_data[:,0], all_store_data[:,2], color='#DA9124')
    ax[0].set_xlabel('epochs', **axis_font)
    ax[0].set_ylabel('loss', **axis_font)
    ax[0].set_title('train / test loss', **title_font)
    ax[0].legend(['train loss', 'test loss'], prop=font_prop, numpoints=1)
    ax[0].hold(False)

    # subplot 2
    ax[1].plot(all_store_data[:,0], all_store_data[:,1], color='#2A6F92', label='test loss')
    ax[1].set_xlabel('epochs', **axis_font)
    ax[1].set_title('train loss', **title_font)


    if report == True:
        plt.subplots_adjust(wspace = wspace, bottom = bottom, hspace  =  hspace)
        w = int(25 / 2.5)  # 30
        h = int(8 / 2.5)  # 10
        fig.set_size_inches(w,h)
        fig.savefig('report_plot.png', dpi=85)


    # adjust image
    w = int(30 / 2.5)  # 30
    h = int(10 / 2.5)  # 10
    fig.set_size_inches(w,h)

    font_prop = font_manager.FontProperties(size=12)
    ax[0].legend(['train loss', 'test loss'], prop=font_prop, numpoints=1)

    top    =  0.8
    plt.subplots_adjust(top = top, hspace = hspace)



    # define title
    plot_title = "Loss values of {} | epochs: {} | {}".format(stage, len(all_store_data), criterion)
    fig.suptitle(plot_title, fontsize=16)


    # -------------------- save the plot ------------------------
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    short_title = "loss_{}_{}_{}".format(stage, len(all_store_data), time_data)


    # create empty folder
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')

    if saveJPG == True:
        fig.savefig('./RESULTS/' + short_title + ".jpg")
    if saveSVG == True:
        fig.savefig('./RESULTS/' + short_title + ".svg")
    if savePDF == True:
        fig.savefig('./RESULTS/' + short_title + ".pdf")





# =========================== SAVE AS .MAT ====================================
def save_as_mat(name, input_array, stage, num_epoch):

    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    short_title = "{}_{}_{}".format(stage, num_epoch, time_data)

    # create empty folder
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')

    # save .mat file
    savemat('./RESULTS/' + name + '_' + short_title + '.mat', input_array)


# =========================== SAVE ARRAY ======================================
# save model information to text file.

def save_net_information(stage, name, model, model_specification):

    # file name
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")

    # add title
    temp_model = []
    temp_model.append(name)
    temp_model.append(time_data)

    # space
    for m in range(4):
        temp_model.append('\n')

    # add model
    temp_model.append('------------------------- MODEL STRUCTURE ----------------------')
    temp_model.append('\n')
    temp_model.append('\n')
    list_model = get_value(model)
    for i in range(len(list_model)):
        temp_model.append(str(list_model[i]))
        temp_model.append('\n')
    temp_model.append('\n')

    # space
    for m in range(3):
        temp_model.append('\n')

    # add net specification
    temp_model.append('---------------------- MODEL SPECIFICATION ---------------------')
    temp_model.append('\n')
    for key, value in model_specification.items():
        spec_key = key
        spec_value = value
        temp_model.append(str(spec_key) + ': ' + str(spec_value))
        temp_model.append('\n')
    temp_model.append('\n')

    # create empty folder
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')

    # save file
    file = open('./RESULTS/' + name + '.txt', 'x')
    file.writelines(temp_model)
    file.close()



# ======================= GET VALUES FROM PYTORCH =============================
# function that takes the files from pytorch that are difficult to obtain
# e.g. model, and convert them into list."""

def get_value(data):

    with open('temp_file.txt', 'w') as f:
        print(data, file=f)

    file = open('temp_file.txt', "r")
    lines = list(file)
    file.close()

    os.remove('temp_file.txt')

    return lines


# ======================= SAVE MATRIX =========================================
def save_data(data, name):

    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')
        
    if type(data) == torch.Tensor:
        data = data.cpu().data.numpy()
    else:
        pass

    new_name = './RESULTS/' + name
    np.save(new_name, data)



# ========================= UPLOAD PLOTS ======================================
def plot_output(stage, epochs, test_labels, test_outputs, img_size, report=False, saveJPG=False,
                saveSVG=False, savePDF=False):

    output = test_labels.view(test_labels.size(0), test_labels.size(-2), test_labels.size(-1))
    pred_output = test_outputs.view(test_outputs.size(0), test_outputs.size(-2), test_outputs.size(-1))

    data_numbers = [1, 2, 3, 4, 5]
    cmap = 'jet'
#    https://matplotlib.org/examples/color/colormaps_reference.html

    fig, ax = plt.subplots(figsize=(12,5), ncols=len(data_numbers), nrows=2)

    for i in range(len(data_numbers)):

        # output
        out = output[data_numbers[i], :, :]
        temp_out = (Variable(out).data).cpu().numpy()
        im = ax[0][i].imshow(temp_out.T, extent=(0, img_size, 0, img_size), cmap=cmap, origin='1')
        ax[0][i].set_axis_off()
        ax[0][i].set_title('output {}'.format(data_numbers[i]))
        divider = make_axes_locatable(ax[0][i])
        cax = divider.append_axes("right", size="5%", pad=0.05)
        plt.colorbar(im, cax=cax)
        
        # predicted output
        pred_out = pred_output[data_numbers[i], :, :]
        temp_pred_out = (Variable(pred_out).data).cpu().numpy()
        im = ax[1][i].imshow(temp_pred_out.T, extent=(0, img_size, 0, img_size), cmap=cmap, origin='1')
        ax[1][i].set_axis_off()
        ax[1][i].set_title('predicted output {}'.format(data_numbers[i]))
        divider = make_axes_locatable(ax[1][i])
        cax = divider.append_axes("right", size="5%", pad=0.05)
        plt.colorbar(im, cax=cax)

    if report == True:
        w = int(25 / 2.5)  # 30
        h = int(11 / 2.5)  # 10
        fig.set_size_inches(w,h)
        plt.tight_layout(pad=0.4, w_pad=0.5, h_pad=1.0)
        fig.savefig('report_output.png', dpi=85)

    # adjust image
    w = int(26 / 2.5)  # 30
    h = int(10 / 2.5)  # 10
    fig.set_size_inches(w,h)

    plt.tight_layout(pad=0.4, w_pad=0.5, h_pad=1.0)
    plt.subplots_adjust(top = 0.8)


    # define title
    plot_title = "Output of {} | epochs: {}".format(stage, epochs)
    fig.suptitle(plot_title, fontsize=16)


    # -------------------- save the plot ------------------------
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    short_title = "output_{}_{}_{}".format(stage, epochs, time_data)


    # create empty folder
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')

    if saveJPG == True:
        fig.savefig('./RESULTS/' + short_title + ".jpg")
    if saveSVG == True:
        fig.savefig('./RESULTS/' + short_title + ".svg")
    if savePDF == True:
        fig.savefig('./RESULTS/' + short_title + ".pdf")



# ========================= UPLOAD IMAGES =====================================
def net_specification(stage, dataset_size, input_size, output_size, epochs,
                      mini_batch_size, learning_rate, momentum, loss_function,
                      optimizer, train_test_split, device, min_train_loss,
                      max_train_loss, min_test_loss, max_test_loss, train_total_loss,
                      test_total_loss, overall_run_time, accuracy, img_diff):

    # get optimizer
    optimizer_type = str(type(optimizer))
    optimizer_type = optimizer_type[20:]
    optimizer_type = optimizer_type[:-2]

    # round numbers
#    train_accuracy = np.round(float(train_accuracy), 5)
#    test_accuracy = np.round(float(test_accuracy), 5)

    # get time
    time_min, time_sec = calculate_time(overall_run_time[0], overall_run_time[1])
    all_time = str(time_min) + ' min  ' + str(time_sec) + ' sec'

    model_specification = dict(stage = stage,
                               dataset_size = dataset_size,
                               input_size = input_size,
                               output_size = output_size,
                               epochs = epochs,
                               mini_batch_size = mini_batch_size,
                               learning_rate = learning_rate,
                               momentum = momentum,
                               loss_function = loss_function,
                               optimizer = optimizer_type,
                               train_test_split = train_test_split,
                               device = device,
                               min_train_loss = min_train_loss,
                               max_train_loss = max_train_loss,
                               min_test_loss = min_test_loss,
                               max_test_loss = max_test_loss,
                               train_total_loss = train_total_loss,
                               test_total_loss = test_total_loss,
                               overall_run_time = all_time,
                               accuracy = accuracy,
                               img_diff = img_diff)

    # save as .mat file
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')

    name = 'model_spec'
    save_as_mat(name, model_specification, stage, epochs)


    return model_specification



# ========================== SAVE MODEL =======================================
def save_model(model, name):

    # create empty folder for model
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')

    torch.save(model, './RESULTS/' + name + '.pt')
    torch.save(model.state_dict(), './RESULTS/' + name + '_state.pt')



# ======================= OBTAIN MAX STRESS ===================================
def max_stress(test_labels, test_outputs, img_size):
    
    test_max_stress = np.zeros((len(test_labels), 2))
    for i in range(len(test_labels)):
        temp_test_labels = test_labels[i]
        temp_test_outputs = test_outputs[i]
        label_max = np.amax((temp_test_labels.view(img_size, img_size)).cpu().data.numpy())
        output_max = np.amax((temp_test_outputs.view(img_size, img_size)).cpu().data.numpy())
        
        test_max_stress[i, 0] = label_max
        test_max_stress[i, 1] = output_max
        
    return test_max_stress

    

# ==================== CREATE PLOT FOR MAX STRESS =============================
def plot_max_stress(test_labels, test_outputs, stage, all_store_data,
                    criterion, report=False, saveJPG=False, saveSVG=False, savePDF=False):
    
#    data_len = int(test_labels.size(0))
    data_len = 50
    
    # get max values of stress
    test_max_stress = max_stress(test_labels[:data_len], test_outputs[:data_len], test_labels.size(-1))

    # font definition
    title_font = {'fontname':'Arial', 'size':'16', 'color':'black', 'weight':'normal',
                  'verticalalignment':'bottom'} # Bottom vertical alignment for more space
    axis_font = {'fontname':'Arial', 'size':'14'}
    font_prop = font_manager.FontProperties(size=11)

    # define figure
    fig, ax = plt.subplots(figsize=(12,5), ncols=1, nrows=1)

    left   =  0.1    # the left side of the subplots of the figure
    right  =  1.1    # the right side of the subplots of the figure
    bottom =  0.2    # the bottom of the subplots of the figure
    top    =  0.9    # the top of the subplots of the figure
    wspace =  .3    # the amount of width reserved for blank space between subplots
    hspace =  .1     # the amount of height reserved for white space between subplots

    ax.vlines(x=np.linspace(0,data_len-1, data_len), ymin=0, ymax=test_max_stress[:,0], color='#C6C6C6')
    ax.scatter(np.linspace(0,data_len-1, data_len), test_max_stress[:,0], color='#2A6F92')
    ax.scatter(np.linspace(0,data_len-1, data_len), test_max_stress[:,1], color='#DA9124')
    ax.set_xlabel('number of test data', **axis_font)
    ax.set_ylabel('max stress value [MPa]', **axis_font)
    ax.set_title('max stress {} values for output and predicted output'.format(data_len), **title_font)
    ax.legend(['outputs', 'predicted outputs'], prop=font_prop, numpoints=1)
    ax.hold(False)

    if report == True:
        plt.subplots_adjust(wspace = wspace, bottom = bottom, hspace  =  hspace)
        w = int(25 / 2.5)  # 30
        h = int(8 / 2.5)  # 10
        fig.set_size_inches(w,h)
        fig.savefig('report_maxstress.png', dpi=85)


    # adjust image
    w = int(30 / 2.5)  # 30
    h = int(11 / 2.5)  # 10
    fig.set_size_inches(w,h)

    font_prop = font_manager.FontProperties(size=12)
    ax.legend(['outputs', 'predicted outputs'], prop=font_prop, numpoints=1)

    top    =  0.8
    plt.subplots_adjust(top = top, hspace = hspace)



    # define title
    plot_title = "Max stress of {} | epochs: {} | {}".format(stage, len(all_store_data), criterion)
    fig.suptitle(plot_title, fontsize=16)


    # -------------------- save the plot ------------------------
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    short_title = "maxstress_{}_{}_{}".format(stage, len(all_store_data), time_data)


    # create empty folder
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')

    if saveJPG == True:
        fig.savefig('./RESULTS/' + short_title + ".jpg")
    if saveSVG == True:
        fig.savefig('./RESULTS/' + short_title + ".svg")
    if savePDF == True:
        fig.savefig('./RESULTS/' + short_title + ".pdf")




# ================== OBATAIN STRESS PROPAGATION VALUES ========================
def take_stress_values(Labels):
    
    show_line = False
    
    # =========================== Draw Line ===================================
    # maximum stress point
    Lab = np.copy(Labels)
    coordinates_of_max_stress = np.zeros((len(Lab), 2))
    for num in range(len(Lab)):
        a = Lab[num]
        a = np.squeeze(a)
        i,j = np.unravel_index(a.argmax(), a.shape)
        coordinates_of_max_stress[num, :] = np.array([i, j]).reshape((1,2))
    
    # center point
    center_point = a = np.array([34, 34]).reshape((1,2))
    
    # draw line
    Lab = np.squeeze(np.copy(Labels))
    new_labels = torch.zeros((len(Lab), 64, 64))
    max_value = np.zeros((len(Lab), 1))
    for num in range(len(Lab)):
        max_value[num, 0] = np.amax(np.squeeze(Lab[num, :, :]))
        im = Image.fromarray(Lab[num])
        draw = ImageDraw.Draw(im) 
        draw.line([coordinates_of_max_stress[num,1],coordinates_of_max_stress[num,0],
                   center_point[0,0], center_point[0,1]], fill=max_value[num, 0])
        temp_img = np.array(im)
        img = torch.from_numpy(temp_img)
        img = img.view(1, 64, 64)
        new_labels[num,:,:] = img
    
    
    # ============================ get idexes =================================
    Lab = np.squeeze(np.copy(Labels))
    New_Lab = np.asarray(new_labels)
    all_idx_list = list()
    
    for num in range(len(New_Lab)):
        temp_img = np.squeeze(New_Lab[num, :, :])
        first_value = 0
        for pixel_x in range(64):
            for pixel_y in range(64):
                if np.amax(temp_img[pixel_y, pixel_x]) == max_value[num, 0]:
                    if first_value == 0:
                        idx_list = np.array([pixel_y, pixel_x]).reshape((1,2))
                        first_value = 1
                    else:
                        idx_list = np.append(idx_list, np.array([pixel_y, pixel_x]).reshape((1,2)), axis=0)
                else:
                    pass
#            if np.amax(temp_img[:, pixel_x]) == max_value[num, 0]:
#                row = temp_img[:, pixel_x]
#                i = np.unravel_index(row.argmax(), row.shape)
#                i = int(i[0])
#                if first_value == 0:
#                    idx_list = np.array([int(i), pixel_x]).reshape((1,2))
#                    first_value = 1
#                else:
#                    idx_list = np.append(idx_list, np.array([int(i), pixel_x]).reshape((1,2)), axis=0)
#            else:
#                pass
        # start all indexes from node
        if (idx_list[0,0]) == 34 and (idx_list[0,1] == 34):
            temp_idx_list = np.copy(idx_list)
            idx_list = np.flipud(temp_idx_list)
        else:
            pass
        
#        if num%1 == 0:
#            print('progress: ', str(num), '/', str(len(New_Lab)))
        
        all_idx_list.append(idx_list)
        
    
    
    # ============================= create plot ===============================
    all_stress_propagation = list()
    labels = np.squeeze(Labels)
    
    for num in range(len(New_Lab)):
        indexes = all_idx_list[num]
        temp_stress_prop = np.zeros((1, len(all_idx_list[num])))
        for idx in range(len(all_idx_list[num])):
            temp_stress = labels[num, indexes[idx,0], indexes[idx,1]]
            temp_stress_prop[0, idx] = temp_stress
            temp_stress_prop = np.sort(temp_stress_prop)
            temp_stress_prop = np.flip(temp_stress_prop, axis=1)
#            if temp_stress_prop[0, 0] < temp_stress_prop[0, len(temp_stress_prop)]:
#                temp_stress_prop = np.flip(temp_stress_prop, axis=1)
#            else:
#                pass
        all_stress_propagation.append(temp_stress_prop)
            
        
    # ============================ Visualize ==================================  
    if show_line == True:

        vis_out = new_labels
        
        num_1 = 0
        num_2 = 1
        num_3 = 2
        num_4 = 3
        num_5 = 4
        num_6 = 5
        num_7 = 6
        num_8 = 7
        
        img_0 = vis_out[num_1, :, :]
        img_1 = vis_out[num_2, :, :]
        img_2 = vis_out[num_3, :, :]
        img_3 = vis_out[num_4, :, :]
        img_4 = vis_out[num_5, :, :]
        img_5 = vis_out[num_6, :, :]
        img_6 = vis_out[num_7, :, :]
        img_7 = vis_out[num_8, :, :]
        
        grid_z0 = (Variable(img_0).data).cpu().numpy()
        grid_z1 = (Variable(img_1).data).cpu().numpy()
        grid_z2 = (Variable(img_2).data).cpu().numpy()
        grid_z3 = (Variable(img_3).data).cpu().numpy()
        grid_z4 = (Variable(img_4).data).cpu().numpy()
        grid_z5 = (Variable(img_5).data).cpu().numpy()
        grid_z6 = (Variable(img_6).data).cpu().numpy()
        grid_z7 = (Variable(img_7).data).cpu().numpy()
        
        plt.subplot(241)
        plt.imshow(grid_z0.T, extent=(0, 64, 0, 64), origin='1')
        plt.subplot(242)
        plt.imshow(grid_z1.T, extent=(0, 64, 0, 64), origin='2')
        plt.subplot(243)
        plt.imshow(grid_z2.T, extent=(0, 64, 0, 64), origin='3')
        plt.subplot(244)
        plt.imshow(grid_z3.T, extent=(0, 64, 0, 64), origin='4')
        plt.subplot(245)
        plt.imshow(grid_z4.T, extent=(0, 64, 0, 64), origin='1')
        plt.subplot(246)
        plt.imshow(grid_z5.T, extent=(0, 64, 0, 64), origin='2')
        plt.subplot(247)
        plt.imshow(grid_z6.T, extent=(0, 64, 0, 64), origin='3')
        plt.subplot(248)
        plt.imshow(grid_z7.T, extent=(0, 64, 0, 64), origin='4')
        
        
        #plt.show()
        plt.savefig('./draw_line.pdf')
    
    return all_stress_propagation



# ==================== STRESS PROPAGATION PLOT ================================
def draw_stress_propagation_plot(Labels, Outputs, stage, all_store_data,
                                 criterion, report=False, saveJPG=False,
                                 saveSVG=False, savePDF=False):
    
    lab_prop = take_stress_values((Labels.data).cpu().numpy())
    out_prop = take_stress_values((Outputs.data).cpu().numpy())
    
    ncols = 2
    nrows = 2
    
    # font definition
    title_font = {'fontname':'Arial', 'size':'16', 'color':'black', 'weight':'normal',
                  'verticalalignment':'bottom'} # Bottom vertical alignment for more space
    axis_font = {'fontname':'Arial', 'size':'14'}
    font_prop = font_manager.FontProperties(size=11)

    # define figure
    fig, ax = plt.subplots(figsize=(12,10), ncols=ncols, nrows=nrows)

    left   =  0.1    # the left side of the subplots of the figure
    right  =  1.1    # the right side of the subplots of the figure
    bottom =  0.1    # the bottom of the subplots of the figure
    top    =  0.9    # the top of the subplots of the figure
    wspace =  .3    # the amount of width reserved for blank space between subplots
    hspace =  .8     # the amount of height reserved for white space between subplots

    index = 0
    for i in range(nrows):
        for j in range(ncols):
            lab_matrix = np.squeeze(lab_prop[index])
            out_matrix = np.squeeze(out_prop[index])
            ax[i][j].hold(True)
            ax[i][j].plot(np.linspace(0,len(lab_matrix)-1, len(lab_matrix)), lab_matrix, color='#2A6F92')
            ax[i][j].plot(np.linspace(0,len(out_matrix)-1, len(out_matrix)), out_matrix, color='#DA9124')
            ax[i][j].set_xlabel('distance from node', **axis_font)
            ax[i][j].set_ylabel('stress [MPa]', **axis_font)
            ax[i][j].set_title('stress propagation {}'.format(index), **title_font)
            ax[i][j].legend(['output', 'Predicted output'], prop=font_prop, numpoints=1)
            ax[i][j].hold(False)
            index += 1

#    # subplot 2
#    ax[1].plot(all_store_data[:,0], all_store_data[:,1], color='#2A6F92', label='test loss')
#    ax[1].set_xlabel('distance from node', **axis_font)
#    ax[1].set_title('stress propagation (predicted output)', **title_font)


    if report == True:
        plt.subplots_adjust(wspace = wspace, bottom = bottom, hspace  =  hspace, top = top)
        w = int(25 / 2.5)  # 30
        h = int(16 / 2.5)  # 10
        fig.set_size_inches(w,h)
        fig.savefig('report_stressprop.png', dpi=85)


    # adjust image
    w = int(30 / 2.5)  # 30
    h = int(25 / 2.5)  # 10
    fig.set_size_inches(w,h)

#    font_prop = font_manager.FontProperties(size=12)
#    ax[0].legend(['output', 'Predicted output'], prop=font_prop, numpoints=1)

    hspace =  .5
    top    =  0.9
    plt.subplots_adjust(top = top, hspace = hspace)



    # define title
    plot_title = "Stress Propagation of {} | epochs: {} | {}".format(stage, len(all_store_data), criterion)
    fig.suptitle(plot_title, fontsize=16)


    # -------------------- save the plot ------------------------
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    short_title = "stressprop{}_{}_{}".format(stage, len(all_store_data), time_data)


    # create empty folder
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')

    if saveJPG == True:
        fig.savefig('./RESULTS/' + short_title + ".jpg")
    if saveSVG == True:
        fig.savefig('./RESULTS/' + short_title + ".svg")
    if savePDF == True:
        fig.savefig('./RESULTS/' + short_title + ".pdf")




# ========================= CREATE REPORT =====================================
def create_report(model, stage, model_specification):

    # ----------------------- Read File ---------------------------------------
    # Load in the workbook
    wb = load_workbook('./template/template_test.xlsx')

    # grab the active worksheet
    ws = wb.active

    # file name
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    file_name = 'REPORT_' + stage + '_' + time_data

    # save .txt file
    save_net_information(stage, file_name, model, model_specification)

    # ----------------------- Upload Data to File -----------------------------
    # get model
    model_description = get_value(model)

    # upload model
    ws['A7'] = "".join(str(x) for x in model_description)

    # upload stage
    ws['D3'] = str(stage)

    # upload save specification
    ws['J8'] = str(model_specification['dataset_size'])
    ws['J9'] = str(model_specification['input_size'])
    ws['J10'] = str(model_specification['output_size'])
    ws['J11'] = str(model_specification['epochs'])
    ws['J12'] = str(model_specification['learning_rate'])
    ws['J13'] = str(model_specification['momentum'])
    ws['J14'] = str(model_specification['loss_function'])
    ws['J15'] = str(model_specification['optimizer'])
    ws['J16'] = str(model_specification['train_test_split'])
    ws['J17'] = str(model_specification['device'])
    ws['J18'] = str(model_specification['mini_batch_size'])

    ws['N8'] = str(model_specification['min_train_loss'])
    ws['N9'] = str(model_specification['max_train_loss'])
    ws['N10'] = str(model_specification['min_test_loss'])
    ws['N11'] = str(model_specification['max_test_loss'])
    ws['N12'] = str(model_specification['train_total_loss'])
    ws['N13'] = str(model_specification['test_total_loss'])
    ws['N17'] = str(model_specification['overall_run_time'])
#    ws['N18'] = str(model_specification['accuracy'])
    ws['N18'] = str(model_specification['img_diff'])

    # upload architecture design graph
    img = drawing.image.Image('./template/' + stage[:9] + '.png')
    ws.add_image(img, 'H21')
    
    # uplaod plot
    img = drawing.image.Image('report_plot.png')
    ws.add_image(img, 'A35')

    # uplaod output
    img = drawing.image.Image('report_output.png')
    ws.add_image(img, 'A49')

    # ------------------------Save File ---------------------------------------
    # save as .xlsx file
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')
    
    wb.save('./RESULTS/' + file_name + '.xlsx')
    
#    os.system('cd RESULTS && convert ' + file_name + '.xlsx ' + file_name + '.pdf')
#    os.system('mv  -v ./RESULTS/_dirname/magick-*.pdf ./RESULTS/' + file_name + '.pdf')
#    os.system('cd RESULTS && rmdir _dirname && cd .. && rm report_plot.png && rm report_output.png')
    
    
    
# ====================== CREATE REPORT - PAGE 2 ===============================
def create_report_page2(model, stage, model_specification):

    # ----------------------- Read File ---------------------------------------
    # Load in the workbook
    wb = load_workbook('./template/template_test_page2.xlsx')

    # grab the active worksheet
    ws = wb.active

    # file name
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    file_name = 'REPORT_' + stage + '_' + time_data

    # ----------------------- Upload Data to File -----------------------------
    # upload stage
    ws['D3'] = str(stage)

    # upload save specification
#    ws['J8'] = str(model_specification['dataset_size'])
#    ws['J9'] = str(model_specification['input_size'])

    ws['N8'] = str(model_specification['min_train_loss'])
    ws['N9'] = str(model_specification['max_train_loss'])
    ws['N10'] = str(model_specification['min_test_loss'])
    ws['N11'] = str(model_specification['max_test_loss'])
    ws['N12'] = str(model_specification['train_total_loss'])
    ws['N13'] = str(model_specification['test_total_loss'])
    
    ws['N17'] = str(model_specification['overall_run_time'])
    ws['N18'] = str(model_specification['img_diff'])
    
    # uplaod stress propagation mean and standard deviation graph
    img = drawing.image.Image('report_stresspropall.png')
    ws.add_image(img, 'A8')

    # uplaod min/max compare stress plot
    img = drawing.image.Image('report_maxstress.png')
    ws.add_image(img, 'A22')

    # uplaod stress propagation plot
    img = drawing.image.Image('report_stressprop.png')
    ws.add_image(img, 'A38')

    # ------------------------Save File ---------------------------------------
    # save as .xlsx file
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')
    
    wb.save('./RESULTS/' + file_name + '_p2.xlsx')
    
#    os.system('cd RESULTS && convert ' + file_name + '_p2.xlsx ' + file_name + '_p2.pdf')
#    os.system('mv  -v ./RESULTS/_dirname/magick-*.pdf ./RESULTS/' + file_name + '_p2.pdf')
#    os.system('cd RESULTS && rmdir _dirname && cd .. && rm report_maxstress.png && rm report_stressprop.png')
      
    
    

# ======================= PLOT SAMPLE PARAMS AND LABELS =======================
def plot_sample_param_and_label(params, labels):
    
    if type(params) == torch.Tensor:
        pass
    else:
        params = torch.from_numpy(params)
    
    if type(labels) == torch.Tensor:
        pass
    else:
        labels = torch.from_numpy(labels)
    
    # get number
    number = random.randint(0, int(labels.size(0)))
    
    # label
    label_img = labels.view(labels.size(0) * labels.size(1), 64, 64)
    grid_label = (Variable(label_img[number, :, :]).data).cpu().numpy()
    plt.imshow(grid_label.T, extent=(0, 64, 0, 64), origin='1')
    plt.savefig('one.svg')
    
    # params
    print(params[number,0,:,:])  
    
    

    
# ================== CREATE CHANNEL X AND Y FOR INPUT IMAGES ==================
def X_and_Y_channels(Inputs, Labels, img_size):
    
    x = np.arange(-1, 1, 2/img_size)
    y = np.arange(-1, 1, 2/img_size)
    y = np.flipud(y)
    xx, yy = np.meshgrid(x, y)
    
    xx = torch.from_numpy(xx)
    yy = torch.from_numpy(yy)
    xx = xx.view(1, 1, 1, 64, 64)
    yy = yy.view(1, 1, 1, 64, 64)
    
    Inputs = torch.from_numpy(Inputs)
    Labels = torch.from_numpy(Labels)
    dataset_len = Inputs.size(0)
    
    new_Inputs = torch.zeros((dataset_len, 1, 3, img_size, img_size))
    new_Labels = torch.zeros((dataset_len, 1, 3, img_size, img_size))
    
    for i in range(dataset_len):
        temp_Inputs = Inputs[i]
        temp_Labels = Labels[i]
        temp_Inputs = temp_Inputs.view(1, 1, 1, img_size, img_size)
        temp_Labels = temp_Labels.view(1, 1, 1, img_size, img_size)
        
        con_Inputs = torch.cat((temp_Inputs, xx, yy), 2)
        new_Inputs[i,:,:,:,:] = con_Inputs
        con_Labels = torch.cat((temp_Labels, xx, yy), 2)
        new_Labels[i,:,:,:,:] = con_Labels
        
    return new_Inputs, new_Labels



# ================== MEAN AND CONFIDENCE INTERVAL DATA ========================
def mean_and_CI(Labels, Outputs, stage, all_store_data, criterion, report=False,
                saveJPG=False, saveSVG=False, savePDF=False):
    
    # get important channel
    if Labels.size(-3) == 3:
        Labels = Labels[:,0,:,:]
        Outputs = Outputs[:,0,:,:]
    elif Labels.size(-3) == 1:
        pass
    else:
        print('there is error in the number of channels')
        
    # get the stress
    lab_prop = take_stress_values((Labels.data).cpu().numpy())
    out_prop = take_stress_values((Outputs.data).cpu().numpy())
    
    min_val = 100
    for i in range(len(out_prop)):
        temp_min_1 = np.shape(lab_prop[i])[1]
        temp_min_2 = np.shape(out_prop[i])[1]
        temp_min_val = np.min((temp_min_1, temp_min_2))
        min_val = np.min((temp_min_val, min_val))
    

    labels_mean = np.zeros((min_val))
    outputs_mean = np.zeros((min_val))
    labels_up = np.zeros((min_val))
    labels_down = np.zeros((min_val))
    outputs_up = np.zeros((min_val))
    outputs_down = np.zeros((min_val)) 
        
    for i in range(min_val):
        
        temp_labels = np.zeros((Labels.size(0)))
        temp_outputs = np.zeros((Outputs.size(0)))
        
        for j in range(Labels.size(0)):
            
            temp_lab = np.squeeze(lab_prop[j])
            temp_out = np.squeeze(out_prop[j])
            temp_labels[j] = float(temp_lab[i])
            temp_outputs[j] = float(temp_out[i])
            
        # get mean
        labels_mean[i] = np.mean((temp_labels))
        outputs_mean[i] = np.mean((temp_outputs))
        
        # get upper and lower boundary for confidence interval
        labels_up[i] = np.max((temp_labels))
        labels_down[i] = np.min((temp_labels))
        outputs_up[i] = np.max((temp_outputs))
        outputs_down[i] = np.min((temp_outputs)) 

    
    # font definition
    title_font = {'fontname':'Arial', 'size':'16', 'color':'black', 'weight':'normal',
                  'verticalalignment':'bottom'} # Bottom vertical alignment for more space
    axis_font = {'fontname':'Arial', 'size':'14'}
    font_prop = font_manager.FontProperties(size=11)

    # define figure
    fig, ax = plt.subplots(figsize=(12,5), ncols=1, nrows=1)

    left   =  0.1    # the left side of the subplots of the figure
    right  =  1.1    # the right side of the subplots of the figure
    bottom =  0.2    # the bottom of the subplots of the figure
    top    =  0.9    # the top of the subplots of the figure
    wspace =  .3    # the amount of width reserved for blank space between subplots
    hspace =  .1     # the amount of height reserved for white space between subplots

    # labels
    ax.fill_between(range(min_val), labels_up, labels_down, color='#85D1F7', alpha=.5)
    ax.plot(labels_mean, '#2A6F92')
    # outputs
    ax.fill_between(range(min_val), outputs_up, outputs_down, color='#FDD9A2', alpha=.5)
    ax.plot(outputs_mean, '#DA9124')

    ax.set_xlabel('distance from node', **axis_font)
    ax.set_ylabel('stress [MPa]', **axis_font)
    ax.set_title('stress propagation', **title_font)
    ax.legend(['outputs', 'predicted outputs'], prop=font_prop, numpoints=1)
    ax.hold(False)


    if report == True:
        plt.subplots_adjust(wspace = wspace, bottom = bottom, hspace  =  hspace)
        w = int(14 / 2.5)  # 30
        h = int(8 / 2.5)  # 10
        fig.set_size_inches(w,h)
        fig.savefig('report_stresspropall.png', dpi=85)


    # adjust image
    w = int(30 / 2.5)  # 30
    h = int(14 / 2.5)  # 10
    fig.set_size_inches(w,h)

    font_prop = font_manager.FontProperties(size=12)
    ax.legend(['outputs', 'predicted outputs'], prop=font_prop, numpoints=1)

    top    =  0.8
    plt.subplots_adjust(top = top, hspace = hspace)



    # define title
    plot_title = "Stress propagation (Mean and 95% conf interval) of {} | epochs: {} | {}".format(stage, len(all_store_data), criterion)
    fig.suptitle(plot_title, fontsize=16)


    # -------------------- save the plot ------------------------
    time_data = datetime.datetime.now().strftime("%y-%m-%d-%H-%M")
    short_title = "stresspropall_{}_{}_{}".format(stage, len(all_store_data), time_data)


    # create empty folder
    if not os.path.exists('./RESULTS'):
        os.mkdir('./RESULTS')

    if saveJPG == True:
        fig.savefig('./RESULTS/' + short_title + ".jpg")
    if saveSVG == True:
        fig.savefig('./RESULTS/' + short_title + ".svg")
    if savePDF == True:
        fig.savefig('./RESULTS/' + short_title + ".pdf")
        

# ================== SAVE BEST MODEL AND STATE_DICT ===========================
def get_best_model(model, name, best_test_loss_value, test_loss_value):
    
    if test_loss_value < best_test_loss_value:
        best_test_loss_value = test_loss_value
    else:
        pass
    
    save_model(model, 'best_' + name)
    
    return best_test_loss_value


# =================== GET LOSS AND OUTPUT DATASET =============================
def get_loss_and_output(model, inputs, labels, mini_batch_size, criterion, device):
    
    num_channels = inputs.size(-3)

    # prepare to size the size of the tensor
    inputs = inputs.view(inputs.size(0), num_channels, inputs.size(-2), inputs.size(-1))
    labels = labels.view(labels.size(0), num_channels, labels.size(-2), labels.size(-1))

    # obtain prediction
    n = 0
    loss = 0
    dataset = inputs.size(0)
    step = mini_batch_size
    
    outputs = np.zeros((dataset, labels.size(1), labels.size(2), labels.size(3)))

    for i in range(0, dataset, step):
        input = inputs[i:i+step, :, :, :]  # train_input.size --> (:, 1, 64, 64)
        label = labels[i:i+step, :, :, :]  # train_labels.size --> (:, 1, 64, 64)
        output = model(input)

        # calculate test_loss
        loss += criterion(output, label).cpu().data.numpy()

        outputs[i:i+step, :, :, :] = output.cpu().data.numpy()
        n += 1

#    total_loss = (loss / n) / dataset**2
    total_loss = loss / dataset**2
    total_loss = np.round(float(total_loss), 5)
    
    outputs_torch = torch.from_numpy(outputs)

    return total_loss, outputs_torch
